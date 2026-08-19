import argparse
import subprocess
from pathlib import Path

from openpyxl import load_workbook


def sql_quote(value):
    if value is None or value == "":
        return "NULL"

    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def main():
    parser = argparse.ArgumentParser(description="Import Sheriff Training officers from an Excel workbook.")
    parser.add_argument("workbook", help="Path to workbook with First Name, Last Name, Email, Division columns.")
    parser.add_argument("--mysql", default=r"C:\wamp64\bin\mysql\mysql9.1.0\bin\mysql.exe")
    parser.add_argument("--database", default="jc_data_portal")
    parser.add_argument("--user", default="root")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb.active

    rows = []
    seen = set()
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        first, last, email, division = [
            str(value).strip() if value is not None else ""
            for value in (list(row[:4]) + [None, None, None, None])[:4]
        ]

        if not (first or last or email or division):
            continue

        if not first or not last:
            skipped += 1
            continue

        key = email.lower() if email else (first.lower(), last.lower())
        if key in seen:
            skipped += 1
            continue

        seen.add(key)
        rows.append((first, last, email, division))

    statements = ["START TRANSACTION;"]
    for division in sorted({division for _, _, _, division in rows if division}):
        statements.append(
            "INSERT INTO sheriff_training_divisions (name, sort_order, is_active) "
            f"VALUES ({sql_quote(division)}, 100, 1) "
            "ON DUPLICATE KEY UPDATE is_active = 1;"
        )

    for first, last, email, division in rows:
        statements.append("SET @existing_id := NULL;")
        if email:
            statements.append(
                "SELECT @existing_id := id "
                "FROM sheriff_training_officers "
                f"WHERE LOWER(email) = {sql_quote(email.lower())} "
                "LIMIT 1;"
            )
        statements.append(
            "SELECT @existing_id := COALESCE(@existing_id, ("
            "SELECT id FROM sheriff_training_officers "
            f"WHERE LOWER(first_name) = {sql_quote(first.lower())} "
            f"AND LOWER(last_name) = {sql_quote(last.lower())} "
            "LIMIT 1));"
        )
        statements.append(
            "INSERT INTO sheriff_training_officers "
            "(first_name, last_name, email, rank_title, division, is_active) "
            f"SELECT {sql_quote(first)}, {sql_quote(last)}, {sql_quote(email)}, NULL, {sql_quote(division)}, 1 "
            "WHERE @existing_id IS NULL;"
        )
        statements.append(
            "UPDATE sheriff_training_officers "
            f"SET first_name = {sql_quote(first)}, "
            f"last_name = {sql_quote(last)}, "
            f"email = {sql_quote(email)}, "
            f"division = {sql_quote(division)}, "
            "is_active = 1 "
            "WHERE id = @existing_id;"
        )

    statements.append("COMMIT;")

    subprocess.run(
        [args.mysql, f"-u{args.user}", args.database],
        input="\n".join(statements),
        text=True,
        check=True,
    )

    print(f"Imported or updated {len(rows)} officers.")
    print(f"Skipped {skipped} blank, incomplete, or duplicate rows.")
    print(f"Divisions found: {', '.join(sorted({division for _, _, _, division in rows if division}))}")


if __name__ == "__main__":
    main()
